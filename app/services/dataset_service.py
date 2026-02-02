# services/dataset_service.py
from __future__ import annotations

import io
from datetime import date, datetime
import re
from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl import load_workbook


# =======================
# Result / Metadata
# =======================

@dataclass
class SheetSkipInfo:
    sheet: str
    reason: str
    has_charts: bool = False
    non_empty_cells_sample: int = 0


@dataclass
class DatasetResult:
    df_dict: dict[str, pd.DataFrame] | None
    errors: list[str]
    warnings: list[str]
    skipped: list[SheetSkipInfo]


# =======================
# Normalization helpers
# =======================

def _is_empty_df(df: pd.DataFrame | None) -> bool:
    return df is None or df.empty


def _strip_accents_basic(s: str) -> str:
    table = str.maketrans(
        "áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
        "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC",
    )
    return str(s).translate(table)


def _norm(s: str) -> str:
    s = _strip_accents_basic(str(s).strip().lower())
    s = re.sub(r"\s+", " ", s)
    return s


def _standardize_sheet_key(sheet_name: str) -> str:
    n = _norm(sheet_name)
    if "na semanal" in n:
        return "NA Semanal"
    if "in situ" in n or "insitu" in n:
        return "In Situ"
    if "bombeado" in n:
        return "Volume Bombeado"
    if "infiltrado" in n:
        return "Volume Infiltrado"
    return sheet_name


def _classify_in_situ_df(df: pd.DataFrame) -> str:
    if df is None or df.empty:
        return "pontos"
    point_col = None
    for col in df.columns:
        ncol = _norm_colname(col)
        if "poco" in ncol or "ponto" in ncol:
            point_col = col
            break
    if point_col is None:
        return "geral"
    points = (
        df[point_col]
        .dropna()
        .astype(str)
        .map(lambda x: _norm(x))
        .unique()
        .tolist()
    )
    points = [p for p in points if p]
    if not points:
        return "geral"
    if len(points) == 1:
        if points[0] in {"geral", "saida", "acumulado", "media", "total"}:
            return "geral"
    general_names = {"entrada", "saida", "geral", "acumulado", "media", "total"}
    if all(p in general_names for p in points):
        return "geral"
    return "pontos"


def _norm_colname(c: str) -> str:
    return _norm(c).replace("/", " ").replace("-", " ").replace("_", " ")


def is_chart_sheet_name(name: str) -> bool:
    n = _norm(name)
    return (
        n.startswith("grafico")
        or " grafico" in n
        or "chart" in n
        or n.startswith("gráfico")
        or " gráfico" in n
    )


def canonical_sheet_name(name: str) -> str:
    """
    Produz um "assunto" da aba para relacionar gráficos com abas tabulares.
    Ex:
      - "Gráfico Volume Bombeado" -> "bombeado"
      - "Volume Bombeado" -> "bombeado"
      - "Gráfico Infiltrado" -> "infiltrado"
      - "Volume Infiltrado" -> "infiltrado"
    """
    n = _norm(name)

    stop_words = [
        "grafico", "gráfico",
        "grafico de", "grafico do", "grafico da",
        "chart", "dashboard", "resumo", "capa",
        "volume", "volumes",
    ]
    for w in stop_words:
        n = n.replace(w, " ")

    n = re.sub(r"[^a-z0-9]+", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


# =======================
# Remove mês/ano se existe data completa
# =======================

def _pick_best_date_col(df: pd.DataFrame) -> str | None:
    if _is_empty_df(df):
        return None

    cols = list(df.columns)
    strong = []
    for c in cols:
        n = _norm_colname(str(c))
        if any(k in n for k in ["data", "date", "timestamp", "datetime"]):
            strong.append(c)

    candidates = strong + [c for c in cols if c not in strong]

    best_col = None
    best_score = 0.0

    for c in candidates:
        s = df[c]
        if pd.api.types.is_datetime64_any_dtype(s):
            score = s.notna().mean()
        else:
            conv = pd.to_datetime(s, errors="coerce", dayfirst=True)
            score = conv.notna().mean()

        if score > best_score and score >= 0.50:
            best_score = score
            best_col = c

        if best_col is not None and best_score >= 0.85 and c in strong:
            break

    return best_col


def _find_redundant_month_year_cols(df: pd.DataFrame) -> list[str]:
    redundants = []
    for c in df.columns:
        n = _norm_colname(str(c))

        is_month = (n == "mes") or (n == "mês") or (n == "month") or n.startswith("mes ") or n.endswith(" mes") or " month" in n
        is_year = (n == "ano") or (n == "year") or n.startswith("ano ") or n.endswith(" ano") or " year" in n
        is_month_year_combo = (("mes" in n or "mês" in n or "month" in n) and ("ano" in n or "year" in n))

        if is_month or is_year or is_month_year_combo:
            redundants.append(c)

    return redundants


def drop_month_year_if_full_date_exists(df: pd.DataFrame) -> pd.DataFrame:
    if _is_empty_df(df):
        return df

    date_col = _pick_best_date_col(df)
    if not date_col:
        return df

    # garante datetime
    try:
        if not pd.api.types.is_datetime64_any_dtype(df[date_col]):
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
    except Exception:
        return df

    # se a data for válida em boa parte das linhas, aplica regra
    valid_rate = df[date_col].notna().mean()
    if valid_rate < 0.50:
        return df

    redundants = _find_redundant_month_year_cols(df)
    redundants = [c for c in redundants if c != date_col]

    if redundants:
        df = df.drop(columns=redundants, errors="ignore")

    return df


# =======================
# Cleaning
# =======================

def clean_basic(df: pd.DataFrame) -> pd.DataFrame:
    if _is_empty_df(df):
        return df

    df = df.copy()

    # remove colunas "Unnamed:*"
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed:.*")]
    # remove colunas/linhas totalmente vazias (isso remove "Coluna1" se for toda None)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")

    # normaliza nomes de colunas
    new_cols = []
    for c in df.columns:
        c_str = str(c).replace("\n", " ").replace("\r", " ")
        c_str = re.sub(r"\s+", " ", c_str).strip()
        new_cols.append(c_str)
    df.columns = new_cols

    # tenta converter datas
    for col in df.columns:
        ncol = _norm_colname(col)
        if "data" in ncol or "date" in ncol or "timestamp" in ncol:
            try:
                df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
            except Exception:
                pass

    # tenta converter números com vírgula quando parece numérico
    for col in df.columns:
        if df[col].dtype == object:
            sample = df[col].dropna().astype(str).head(60)
            if len(sample) == 0:
                continue

            looks_numeric = sample.str.match(
                r"^\s*[<>]?\s*[-+]?\d{1,3}(\.\d{3})*(,\d+)?\s*$|^\s*[<>]?\s*[-+]?\d+([.,]\d+)?\s*$"
            ).mean()

            if looks_numeric >= 0.6:
                s = df[col].astype(str).str.strip()
                s = s.str.replace(" ", "", regex=False)
                s_num = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
                s_num = s_num.str.replace("<", "", regex=False).str.replace(">", "", regex=False)
                df[col] = pd.to_numeric(s_num, errors="coerce")

    # ✅ remove Mês/Ano se existir Data completa válida
    df = drop_month_year_if_full_date_exists(df)

    # ✅ remove colunas que ficaram 100% vazias depois das conversões
    df = df.dropna(axis=1, how="all")

    return df


# =======================
# Sheet detection (charts vs table)
# =======================

def _looks_like_na_param(value: object) -> str | None:
    if value is None:
        return None
    n = _norm(str(value))
    n = n.replace("(", " ").replace(")", " ")
    n = re.sub(r"\s+", " ", n).strip()
    if n in ("na m", "na"):
        return "NA (m)"
    if n in ("no m", "no"):
        return "NO (m)"
    if n in ("fl m", "fl"):
        return "FL (m)"
    return None


def _to_float_maybe_units(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s or s == "-":
        return None

    s_low = _norm(s)
    nums = re.findall(r"[-+]?\d+(?:[.,]\d+)?", s_low)
    if not nums:
        return None

    num = nums[0].replace(".", "").replace(",", ".") if "," in nums[0] else nums[0]
    try:
        v = float(num)
    except Exception:
        return None

    # se vier em mS, converte para uS
    if "ms" in s_low:
        v *= 1000.0
    return v


def _extract_in_situ_ws(ws) -> pd.DataFrame | None:
    records: list[dict[str, object]] = []
    current_point = None
    current_date = None
    param_cols: dict[int, str] = {}

    max_row = ws.max_row
    max_col = ws.max_column

    def _is_date(v: object) -> bool:
        return isinstance(v, (pd.Timestamp, datetime, date))

    for r in range(1, max_row + 1):
        b = ws.cell(r, 2).value  # col B
        c = ws.cell(r, 3).value  # col C

        if isinstance(b, str) and _norm(b) == "data" and isinstance(c, str) and c.strip():
            current_point = str(c).strip().upper()
            param_cols = {}
            continue

        if isinstance(c, str) and _norm(c) == "ph":
            for col_idx in range(3, max_col + 1):
                pname = ws.cell(r, col_idx).value
                if pname is None:
                    continue
                pname = str(pname).strip()
                if pname:
                    param_cols[col_idx] = pname
            continue

        if isinstance(c, str) and _norm(c) == "saida":
            current_point = "SAIDA"
            param_cols = {}
            continue

        if param_cols:
            if _is_date(b):
                current_date = b

            if current_date is None or current_point is None:
                continue

            rec = {"Data": pd.to_datetime(current_date), "Ponto": current_point}
            has_any = False
            for col_idx, pname in param_cols.items():
                v = ws.cell(r, col_idx).value
                vnum = _to_float_maybe_units(v)
                if vnum is not None:
                    has_any = True
                rec[pname] = vnum

            if has_any:
                records.append(rec)

    if not records:
        return None

    df = pd.DataFrame.from_records(records)
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["Data"])
    df["Ponto"] = df["Ponto"].astype(str).str.strip().str.upper()
    return df


def _is_na_semanal_sheet(ws) -> bool:
    try:
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    except Exception:
        return False

    if not row1 or not row2:
        return False

    first = _norm(str(row1[0])) if row1[0] is not None else ""
    if "poco" not in first and "ponto" not in first:
        return False

    params = [
        _looks_like_na_param(v)
        for v in row1[1:10]
        if v is not None
    ]
    has_triplet = {p for p in params if p}
    if not {"NA (m)", "NO (m)", "FL (m)"}.issubset(has_triplet):
        return False

    parsed_dates = 0
    for v in row2[1:10]:
        parsed = pd.to_datetime(v, errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            parsed_dates += 1
    return parsed_dates > 0


def _normalize_na_semanal_value(value: object):
    if value is None or pd.isna(value):
        return pd.NA

    if isinstance(value, (int, float)):
        return value

    s = str(value).strip()
    if s == "":
        return pd.NA

    n = _norm(s)
    if n in {"nm", "n m", "nao medido"}:
        return "Não medido"
    if n in {"nd", "n d", "nao localizado", "nao localiz"}:
        return s

    if re.match(r"^[-+]?\d{1,3}(\.\d{3})*(,\d+)?$|^[-+]?\d+([.,]\d+)?$", s):
        s_num = s.replace(" ", "").replace(".", "").replace(",", ".")
        try:
            return float(s_num)
        except Exception:
            return s

    return s

def _is_dt(v: object) -> bool:
    return isinstance(v, (pd.Timestamp, datetime, date))

def _extract_na_semanal_blocks_ws(ws) -> pd.DataFrame | None:
    """
    Formato do arquivo (2): blocos lado a lado.
    Row1: 'ID do poço', <DATA>, ..., (vazio), 'ID do poço', <DATA>, ...
    Row2: headers: NA (m), NO (m), FL, Status, Observação
    """
    try:
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    except Exception:
        return None

    if not row1 or not row2:
        return None

    max_col = max(len(row1), len(row2))
    row1 = list(row1) + [None] * (max_col - len(row1))
    row2 = list(row2) + [None] * (max_col - len(row2))

    # achar índices onde começa um bloco ("ID do poço")
    starts = []
    for i, v in enumerate(row1):
        if v is None:
            continue
        if _norm(str(v)) in {"id do poco", "id do poço", "poco", "poço"}:
            # precisa ter data em i+1
            if i + 1 < max_col and _is_dt(row1[i + 1]):
                starts.append(i)

    if not starts:
        return None

    # valida se a linha 2 parece ter NA/NO/FL perto do primeiro bloco
    i0 = starts[0]
    h_na = _looks_like_na_param(row2[i0 + 1]) if i0 + 1 < max_col else None
    h_no = _looks_like_na_param(row2[i0 + 2]) if i0 + 2 < max_col else None
    if h_na != "NA (m)" or h_no != "NO (m)":
        # ainda pode ser, mas é um bom sinal de falha
        pass

    records: list[dict[str, object]] = []
    empty_streak = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue
        row_vals = list(row) + [None] * (max_col - len(row))

        # para depois de várias linhas vazias
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        for start in starts:
            poco = row_vals[start] if start < max_col else None
            if poco is None or str(poco).strip() == "":
                continue
            poco = str(poco).strip()

            dt = row1[start + 1] if start + 1 < max_col else None
            dt = pd.to_datetime(dt, errors="coerce", dayfirst=True)
            if pd.isna(dt):
                continue

            na_v = _normalize_na_semanal_value(row_vals[start + 1])  # NA (m)
            no_v = _normalize_na_semanal_value(row_vals[start + 2])  # NO (m)
            fl_v = _normalize_na_semanal_value(row_vals[start + 3])  # FL
            stt  = row_vals[start + 4] if start + 4 < max_col else None
            obs  = row_vals[start + 5] if start + 5 < max_col else None

            rec = {
                "Poco": poco,
                "Data": dt,
                "NA (m)": na_v,
                "NO (m)": no_v,
                "FL (m)": fl_v,
            }
            if stt not in (None, "-", ""):
                rec["Status"] = str(stt).strip()
            if obs not in (None, "-", ""):
                rec["Observação"] = str(obs).strip()

            records.append(rec)

    if not records:
        return None

    df = pd.DataFrame.from_records(records)
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["Data", "Poco"])
    df = df.sort_values(["Poco", "Data"]).reset_index(drop=True)
    return df


def _extract_na_semanal_ws(ws) -> pd.DataFrame | None:
    # ✅ 1) tenta formato BLOCO (arquivo (2))
    df_block = _extract_na_semanal_blocks_ws(ws)
    if df_block is not None and not df_block.empty:
        return df_block

    # ✅ 2) fallback: formato MATRIZ (arquivo (8)) - seu código atual abaixo
    try:
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    except Exception:
        return None

    max_col = max(len(row1), len(row2))
    row1 = list(row1) + [None] * (max_col - len(row1))
    row2 = list(row2) + [None] * (max_col - len(row2))

    data_cols: list[tuple[int, str, pd.Timestamp]] = []
    last_date = None
    for idx in range(1, max_col):
        param = _looks_like_na_param(row1[idx])
        if not param:
            continue

        raw_date = row2[idx]
        parsed_date = pd.to_datetime(raw_date, errors="coerce", dayfirst=True)
        if pd.notna(parsed_date):
            last_date = parsed_date
        if last_date is None:
            continue

        data_cols.append((idx, param, last_date))

    if not data_cols:
        return None

    records: list[dict[str, object]] = []
    empty_streak = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue
        row_vals = list(row) + [None] * (max_col - len(row))
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        poco = row_vals[0]
        if poco is None or str(poco).strip() == "":
            continue
        poco = str(poco).strip()

        by_date: dict[pd.Timestamp, dict[str, object]] = {}
        for idx, param, date_val in data_cols:
            raw_val = row_vals[idx] if idx < len(row_vals) else None
            val = _normalize_na_semanal_value(raw_val)
            if date_val not in by_date:
                by_date[date_val] = {"Poco": poco, "Data": date_val}
            by_date[date_val][param] = val

        for date_val in sorted(by_date.keys()):
            rec = by_date[date_val]
            if any(k in rec for k in ("NA (m)", "NO (m)", "FL (m)")):
                records.append(rec)

    if not records:
        return None

    return pd.DataFrame.from_records(records)


def _sheet_has_charts(ws) -> bool:
    try:
        return hasattr(ws, "_charts") and ws._charts and len(ws._charts) > 0
    except Exception:
        return False


def _count_non_empty_cells_sample(ws, max_rows: int = 120, max_cols: int = 40) -> int:
    non_empty = 0
    try:
        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            for v in row:
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                non_empty += 1
    except Exception:
        return 0
    return non_empty


def _extract_table_from_ws(
    ws,
    header_search_rows: int = 40,
    max_cols: int = 80,
    max_data_rows: int = 6000,
) -> pd.DataFrame | None:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_data_rows, max_col=max_cols, values_only=True):
        rows.append(list(row))

    def row_nonempty_count(r: list[Any]) -> int:
        c = 0
        for v in r:
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            c += 1
        return c

    if not rows:
        return None

    search = rows[:header_search_rows]
    counts = [row_nonempty_count(r) for r in search]
    if not counts or max(counts) < 2:
        return None

    header_idx = int(max(range(len(counts)), key=lambda i: counts[i]))
    header = rows[header_idx]
    data_rows = rows[header_idx + 1 :]

    last_col = -1
    for i, v in enumerate(header):
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        last_col = i

    if last_col < 1:
        return None

    header = header[: last_col + 1]
    data_rows = [r[: last_col + 1] for r in data_rows]

    header = [("" if v is None else str(v).strip()) for v in header]
    if sum(1 for h in header if h != "") < 2:
        return None

    seen: dict[str, int] = {}
    fixed: list[str] = []
    for h in header:
        h2 = h if h else "col"
        h2 = re.sub(r"\s+", " ", h2).strip()
        if h2 in seen:
            seen[h2] += 1
            h2 = f"{h2}_{seen[h2]}"
        else:
            seen[h2] = 0
        fixed.append(h2)

    cleaned_rows: list[list[Any]] = []
    empty_streak = 0
    for r in data_rows:
        if row_nonempty_count(r) == 0:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0
        cleaned_rows.append(r)

    if not cleaned_rows:
        return None

    df = pd.DataFrame(cleaned_rows, columns=fixed)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    if df.empty or len(df.columns) < 2:
        return None

    return df

# =======================
# Main entry
# =======================

def build_dataset_from_excel(uploaded_file) -> DatasetResult:
    errors: list[str] = []
    warnings: list[str] = []
    skipped: list[SheetSkipInfo] = []

    try:
        file_bytes = uploaded_file.getvalue()
        bio = io.BytesIO(file_bytes)
    except Exception as e:
        return DatasetResult(df_dict=None, errors=[f"Falha ao ler bytes do upload: {e}"], warnings=[], skipped=[])

    try:
        wb = load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        return DatasetResult(df_dict=None, errors=[f"Falha ao abrir Excel (openpyxl): {e}"], warnings=[], skipped=[])

    sheet_names = list(wb.sheetnames)

    # cria mapa "assunto -> aba tabular" para ignorar "Gráfico X" quando existe a irmã "X"
    canonical_to_data_sheet: dict[str, str] = {}
    for s in sheet_names:
        if not is_chart_sheet_name(s):
            canon = canonical_sheet_name(s)
            if canon:
                canonical_to_data_sheet[canon] = s

    df_dict: dict[str, pd.DataFrame] = {}

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        sn = _norm(sheet_name)
        std_key = _standardize_sheet_key(sheet_name)
        if sn == "na semanal":
            df = _extract_na_semanal_ws(ws)
            if df is None or df.empty:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason="Falha ao interpretar planilha NA semanal.",
                        has_charts=_sheet_has_charts(ws),
                        non_empty_cells_sample=_count_non_empty_cells_sample(ws),
                    )
                )
            else:
                df_dict[std_key] = df
                if "Data" in df.columns:
                    df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
            continue

        if sn == "in situ":
            df = _extract_in_situ_ws(ws)
            if df is None or df.empty:
                try:
                    df = _extract_table_from_ws(ws)
                    if df is not None:
                        df = clean_basic(df)
                except Exception:
                    df = None
            if df is None or df.empty:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason="Falha ao interpretar planilha In Situ.",
                        has_charts=_sheet_has_charts(ws),
                        non_empty_cells_sample=_count_non_empty_cells_sample(ws),
                    )
                )
            else:
                df_dict[std_key] = df
            continue

        if _is_na_semanal_sheet(ws):
            df = _extract_na_semanal_ws(ws)
            if df is None or df.empty:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason="Falha ao interpretar planilha NA semanal.",
                        has_charts=_sheet_has_charts(ws),
                        non_empty_cells_sample=_count_non_empty_cells_sample(ws),
                    )
                )
            else:
                df_dict[std_key] = df
            continue

        has_charts = _sheet_has_charts(ws)
        non_empty_sample = _count_non_empty_cells_sample(ws)

        # 1) ignora "Gráfico ..." se existe a aba irmã tabular
        if is_chart_sheet_name(sheet_name):
            canon = canonical_sheet_name(sheet_name)
            if canon and canon in canonical_to_data_sheet:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason=f"Aba de gráfico ignorada (existe aba tabular correspondente: '{canonical_to_data_sheet[canon]}')",
                        has_charts=has_charts,
                        non_empty_cells_sample=non_empty_sample,
                    )
                )
                continue

        # 2) fallback: tem gráfico e pouco conteúdo tabular
        if has_charts and non_empty_sample < 40:
            skipped.append(
                SheetSkipInfo(
                    sheet=sheet_name,
                    reason="Aba contém gráfico e Não parece ter tabela (amostra com pouco conteúdo).",
                    has_charts=True,
                    non_empty_cells_sample=non_empty_sample,
                )
            )
            continue

        # 3) extrai e limpa tabela
        try:
            df = _extract_table_from_ws(ws)
            if df is None:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason="Não foi possível identificar uma tabela na aba.",
                        has_charts=has_charts,
                        non_empty_cells_sample=non_empty_sample,
                    )
                )
                continue

            df = clean_basic(df)
            if df.empty:
                skipped.append(
                    SheetSkipInfo(
                        sheet=sheet_name,
                        reason="Tabela extraída ficou vazia após limpeza.",
                        has_charts=has_charts,
                        non_empty_cells_sample=non_empty_sample,
                    )
                )
                continue

            df_dict[std_key] = df

        except Exception as e:
            warnings.append(f"Erro ao processar a aba '{sheet_name}': {e}")
            skipped.append(
                SheetSkipInfo(
                    sheet=sheet_name,
                    reason=f"Exceção ao processar: {e}",
                    has_charts=has_charts,
                    non_empty_cells_sample=non_empty_sample,
                )
            )

    if not df_dict:
        errors.append("Nenhuma aba tabular foi encontrada (ou todas foram ignoradas).")
        return DatasetResult(df_dict=None, errors=errors, warnings=warnings, skipped=skipped)

    if skipped:
        warnings.append(f"Foram ignoradas {len(skipped)} abas que Não pareciam tabulares (ex.: gráficos).")

    return DatasetResult(df_dict=df_dict, errors=errors, warnings=warnings, skipped=skipped)


def _merge_df_list(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    if not dfs:
        return pd.DataFrame()
    df = pd.concat(dfs, ignore_index=True, sort=False)
    df = df.drop_duplicates()
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
        df = df.sort_values("Data")
    return df.reset_index(drop=True)


def build_dataset_from_excels(uploaded_files) -> DatasetResult:
    errors: list[str] = []
    warnings: list[str] = []
    skipped: list[SheetSkipInfo] = []

    if not uploaded_files:
        return DatasetResult(df_dict=None, errors=["Nenhum arquivo enviado."], warnings=[], skipped=[])

    grouped: dict[str, list[pd.DataFrame]] = {}
    in_situ_pontos: list[pd.DataFrame] = []
    in_situ_geral: list[pd.DataFrame] = []

    for uploaded_file in uploaded_files:
        result = build_dataset_from_excel(uploaded_file)
        if result.errors:
            name = getattr(uploaded_file, "name", "arquivo")
            errors.extend([f"{name}: {e}" for e in result.errors])
            continue

        warnings.extend(result.warnings)
        skipped.extend(result.skipped)

        if not result.df_dict:
            continue

        for key, df in result.df_dict.items():
            if key == "In Situ":
                if df is None or df.empty:
                    continue
                kind = _classify_in_situ_df(df)
                df_copy = df.copy()
                if "Ponto" not in df_copy.columns:
                    df_copy["Ponto"] = "GERAL"
                if kind == "geral":
                    in_situ_geral.append(df_copy)
                else:
                    in_situ_pontos.append(df_copy)
                continue

            grouped.setdefault(key, []).append(df)

    merged: dict[str, pd.DataFrame] = {}

    for key, dfs in grouped.items():
        merged_df = _merge_df_list(dfs)
        if not merged_df.empty:
            merged[key] = merged_df

    if in_situ_pontos:
        merged["In Situ (Pontos)"] = _merge_df_list(in_situ_pontos)
    if in_situ_geral:
        merged["In Situ (Geral)"] = _merge_df_list(in_situ_geral)

    if not merged:
        errors.append("Nenhuma aba tabular foi encontrada (ou todas foram ignoradas).")
        return DatasetResult(df_dict=None, errors=errors, warnings=warnings, skipped=skipped)

    if skipped:
        warnings.append(f"Foram ignoradas {len(skipped)} abas que NAO pareciam tabulares (ex.: graficos).")

    return DatasetResult(df_dict=merged, errors=errors, warnings=warnings, skipped=skipped)
